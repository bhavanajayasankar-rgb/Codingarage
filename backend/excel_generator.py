import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11, color="0F172A")
MUTED_FONT = Font(name="Calibri", size=10, color="64748B")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="0F172A")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="0F172A")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def format_inr(value: float) -> str:
    return f"\u20B9{value:,.2f}"

def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_body(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER


def _add_title(ws, title, row=1):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def _add_subtitle(ws, text, row):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = MUTED_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_report_excel(report_data: dict, persona: str) -> bytes:
    wb = Workbook()

    # ---------- Summary sheet ----------
    ws = wb.active
    ws.title = "Summary"
    _set_col_widths(ws, [30, 50, 30, 30])

    title = report_data.get("title", f"{persona.capitalize()} Report")
    _add_title(ws, title, 1)
    _add_subtitle(ws, f"{persona.capitalize()} Persona  |  {datetime.now().strftime('%B %d, %Y')}", 2)

    row = 4
    cadence = report_data.get("cadence", "N/A")
    ws.cell(row=row, column=1, value="Cadence").font = SECTION_FONT
    ws.cell(row=row, column=2, value=cadence).font = BODY_FONT
    row += 2

    summary = report_data.get("summary", "") or report_data.get("focus", "")
    if summary:
        ws.cell(row=row, column=1, value="Executive Summary").font = SECTION_FONT
        cell = ws.cell(row=row + 1, column=1, value=summary)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 3, end_column=4)
        row += 5

    metrics = report_data.get("metrics", {})
    if metrics:
        ws.cell(row=row, column=1, value="Key Metrics").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
        ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        _style_header(ws, row, 2)
        row += 1
        for key, val in metrics.items():
            label = key.replace("_", " ").title()
            if isinstance(val, (int, float)):
                if "cost" in key or "savings" in key or "impact" in key or "projected" in key:
                    val_str = format_inr(val)
                else:
                    val_str = f"{val:,}" if val == int(val) else str(val)
            else:
                val_str = str(val)
            ws.cell(row=row, column=1, value=label).font = BODY_FONT
            ws.cell(row=row, column=2, value=val_str).font = BODY_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1
        row += 1

    # ---------- Insights sheet ----------
    ws2 = wb.create_sheet("Insights")
    _set_col_widths(ws2, [10, 80])
    _add_title(ws2, "Key Insights", 1)
    _add_subtitle(ws2, f"{title} - Insights", 2)
    ws2.cell(row=4, column=1, value="#").font = HEADER_FONT
    ws2.cell(row=4, column=2, value="Insight").font = HEADER_FONT
    _style_header(ws2, 4, 2)
    insights = report_data.get("insights", [])
    for i, item in enumerate(insights, 1):
        ws2.cell(row=4 + i, column=1, value=i).font = BODY_FONT
        ws2.cell(row=4 + i, column=2, value=item).font = BODY_FONT
        ws2.cell(row=4 + i, column=1).border = THIN_BORDER
        ws2.cell(row=4 + i, column=2).border = THIN_BORDER

    # ---------- Recommendations sheet ----------
    ws3 = wb.create_sheet("Recommendations")
    _set_col_widths(ws3, [10, 80])
    _add_title(ws3, "Recommendations", 1)
    _add_subtitle(ws3, f"{title} - Recommendations", 2)
    ws3.cell(row=4, column=1, value="#").font = HEADER_FONT
    ws3.cell(row=4, column=2, value="Recommendation").font = HEADER_FONT
    _style_header(ws3, 4, 2)
    recs = report_data.get("recommendations", [])
    for i, item in enumerate(recs, 1):
        ws3.cell(row=4 + i, column=1, value=i).font = BODY_FONT
        ws3.cell(row=4 + i, column=2, value=item).font = BODY_FONT
        ws3.cell(row=4 + i, column=1).border = THIN_BORDER
        ws3.cell(row=4 + i, column=2).border = THIN_BORDER

    # ---------- Action Items sheet ----------
    ws4 = wb.create_sheet("Action Items")
    _set_col_widths(ws4, [10, 80])
    _add_title(ws4, "Action Items", 1)
    _add_subtitle(ws4, f"{title} - Action Items", 2)
    ws4.cell(row=4, column=1, value="#").font = HEADER_FONT
    ws4.cell(row=4, column=2, value="Action Item").font = HEADER_FONT
    _style_header(ws4, 4, 2)
    actions = report_data.get("action_items", [])
    for i, item in enumerate(actions, 1):
        ws4.cell(row=4 + i, column=1, value=i).font = BODY_FONT
        ws4.cell(row=4 + i, column=2, value=item).font = BODY_FONT
        ws4.cell(row=4 + i, column=1).border = THIN_BORDER
        ws4.cell(row=4 + i, column=2).border = THIN_BORDER

    # ---------- Alerts sheet ----------
    ws5 = wb.create_sheet("Alerts")
    _set_col_widths(ws5, [20, 70])
    _add_title(ws5, "Alerts", 1)
    _add_subtitle(ws5, f"{title} - Alerts", 2)
    ws5.cell(row=4, column=1, value="Service").font = HEADER_FONT
    ws5.cell(row=4, column=2, value="Alert").font = HEADER_FONT
    _style_header(ws5, 4, 2)
    alerts = report_data.get("alerts", [])
    for i, alert in enumerate(alerts, 1):
        service = alert.get("service", "Unknown")
        msg = alert.get("alert", "")
        ws5.cell(row=4 + i, column=1, value=service).font = BODY_FONT
        ws5.cell(row=4 + i, column=2, value=msg).font = BODY_FONT
        ws5.cell(row=4 + i, column=1).border = THIN_BORDER
        ws5.cell(row=4 + i, column=2).border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
